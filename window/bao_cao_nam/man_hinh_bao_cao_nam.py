import pandas as pd
import csv
from openpyxl import Workbook, styles
from openpyxl.utils import get_column_letter
from window.bao_cao_nam.ui_man_hinh_bao_cao_nam import Ui_man_hinh_bao_cao_nam
from PyQt6 import QtWidgets

class Man_hinh_bao_cao_nam(QtWidgets.QWidget, Ui_man_hinh_bao_cao_nam):
    def __init__(self, parent = None):
        super().__init__(parent)
        self.setupUi(self)
        self.par = parent
        self.mycursor = self.par.par.conn.cursor()

        # Thêm năm vào combobox năm
        command = """select TenNamHoc
                    from NAMHOC"""
        result = self.mycursor.execute(command).fetchall()
        for row in result:
            self.nam_hoc.addItem(row[0])

        # Chức năng các nút
        self.thoat.clicked.connect(self.Thoat)
        self.tao.clicked.connect(self.BaoCaoNam)
        self.xuat.clicked.connect(self.Xuat)

    def Thoat(self):
        self.par.switch_ManHinhChinh()

    def BaoCaoNam(self):
        self.bao_cao_nam.clearContents()
        self.bao_cao_nam.setRowCount(0)
        if self.nam_hoc.currentIndex() >= 0:
            # Lấy ra mã năm học
            command = f"""select MaNamHoc
                        from NAMHOC
                        where TenNamHoc = N'{self.nam_hoc.currentText()}'"""
            
            ma_nam_hoc = int(self.mycursor.execute(command).fetchall()[0][0])
            
            # Lấy mã giáo viên
            self.MaGiangVien = self.par.par.MaGiangVien

            # Tính tổng số đề thi mà giảng viên ra năm đó
            command = f"""select count(MaDeThi)
                        from DETHI d
                        where d.MaNamHoc = {ma_nam_hoc}
                        and d.MaGiangVien = {self.MaGiangVien}"""
            
            tong_so_de_thi = self.mycursor.execute(command).fetchall()[0][0]

            if tong_so_de_thi == None:
                tong_so_de_thi = 0
            else:
                tong_so_de_thi = int(tong_so_de_thi)

            self.tong_so_de_thi.setText(str(tong_so_de_thi))

            # Tính tổng số bài chấm mà giảng viên chấm năm đó
            command = f"""select sum(kq.TongSoBaiCham)
                        from KETQUACHAMTHI kq, DETHI d
                        where d.MaNamHoc = {ma_nam_hoc}
                        and d.MaGiangVien = {self.MaGiangVien}
                        and kq.MaDeThi = d.MaDeThi"""
            
            tong_so_bai_cham = self.mycursor.execute(command).fetchall()[0][0]

            if tong_so_bai_cham == None:
                tong_so_bai_cham = 0
            else:
                tong_so_bai_cham = int(tong_so_bai_cham)
            
            self.tong_so_bai_cham.setText(str(tong_so_bai_cham))

            # Lấy ra các môn giảng viên đó dạy
            command = f"""select MaMonHoc
                        from CT_GIANGVIEN
                        where MaGiangVien = {self.MaGiangVien}"""
            
            result = self.mycursor.execute(command).fetchall()

            # Thêm vào bảng báo cáo từng môn
            for row, data in enumerate(result):
                    danh_sach_gia_tri = []

                    # Lấy số lượng đề thi của môn đó
                    command = f"""select count(MaDeThi)
                            from DETHI d
                            where d.MaNamHoc = {ma_nam_hoc}
                            and d.MaGiangVien = {self.MaGiangVien}
                            and d.MaMonHoc = {int(data[0])}"""
                    
                    so_luong_de_thi_mon = self.mycursor.execute(command).fetchall()[0][0]

                    if so_luong_de_thi_mon == None:
                        so_luong_de_thi_mon = 0
                    else:
                        so_luong_de_thi_mon = int(so_luong_de_thi_mon)

                    # Lấy số lượng bài chấm của môn đó
                    command = f"""select sum(kq.TongSoBaiCham)
                        from KETQUACHAMTHI kq, DETHI d
                        where d.MaNamHoc = {ma_nam_hoc}
                        and d.MaGiangVien = {self.MaGiangVien}
                        and kq.MaDeThi = d.MaDeThi
                        and d.MaMonHoc = {int(data[0])}"""
            
                    so_luong_bai_cham_mon = self.mycursor.execute(command).fetchall()[0][0]

                    if so_luong_bai_cham_mon == None:
                        so_luong_bai_cham_mon = 0
                    else:
                        so_luong_bai_cham_mon = int(so_luong_bai_cham_mon)

                    # Tính tỉ lệ đề thi
                    try:
                        ti_le_de_thi = float(so_luong_de_thi_mon/tong_so_de_thi)
                    except:
                        ti_le_de_thi = 0.

                    # Tính tỉ lệ bài chấm
                    try:
                        ti_le_bai_cham = float(so_luong_bai_cham_mon/tong_so_bai_cham)
                    except:
                        ti_le_bai_cham = 0.

                    # Lấy tên môn học ra
                    command = f"""select TenMonHoc
                                from MonHoc
                                where MaMonHoc = {int(data[0])}"""
                    ten_mon_hoc = self.mycursor.execute(command).fetchall()[0][0]

                    # Thêm các giá trị vô danh sách
                    danh_sach_gia_tri.append(ten_mon_hoc)
                    danh_sach_gia_tri.append(so_luong_de_thi_mon)
                    danh_sach_gia_tri.append(so_luong_bai_cham_mon)
                    danh_sach_gia_tri.append(ti_le_de_thi)
                    danh_sach_gia_tri.append(ti_le_bai_cham)

                    # Hiển thị các thông tin ra
                    self.bao_cao_nam.insertRow(row)
                    for col, gia_tri in enumerate(danh_sach_gia_tri):
                        self.bao_cao_nam.setItem(row, col, QtWidgets.QTableWidgetItem(str(gia_tri)))
        else:
            msg = QtWidgets.QMessageBox(self)
            msg.setWindowTitle("Thông báo")
            msg.setInformativeText("Vui lòng chọn năm cần báo cáo")
            msg.setFixedSize(300, 150)
            msg.show()

    def Xuat(self):
        with open("Bao_cao.csv", 'w', newline='', encoding= 'utf-8') as file:
            writer = csv.writer(file)
            header_data = [self.bao_cao_nam.horizontalHeaderItem(i).text() for i in range(self.bao_cao_nam.columnCount())]
            writer.writerow(header_data)
            for row in range(self.bao_cao_nam.rowCount()):
                row_data = [self.bao_cao_nam.item(row, col).text() for col in range(self.bao_cao_nam.columnCount())]
                writer.writerow(row_data)
        data = pd.read_csv("Bao_cao.csv")

        df = pd.DataFrame(data)

        wb = Workbook()
        ws = wb.active

        # ws.merge_cells("A1:A3")
        ws.merge_cells("A1:E3")
        ws.merge_cells("A4:E4")

        ws['A1'] = "Báo cáo năm"
        ws['A4'] = f"{self.nam_hoc.currentText()}"
        ws["A1"].alignment = styles.Alignment(horizontal = 'center')
        ws["A4"].alignment = styles.Alignment(horizontal = 'center')
        headers = list(df.columns)
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws[f'{col_letter}7'] = header

        for row_data in df.itertuples(index=False):
            ws.append(list(row_data))

        wb.save('du_lieu.xlsx')

        msg = QtWidgets.QMessageBox(self)
        msg.setWindowTitle("Thông báo")
        msg.setInformativeText("Xuất báo cáo thành công")
        msg.show()

    def DatMacDinh(self):
        # Đặt chính sách kích thước cho MyWidget
        self.setFixedSize(960, 540)
        
        self.nam_hoc.setCurrentIndex(-1)

        self.bao_cao_nam.setColumnCount(5)
        self.bao_cao_nam.setHorizontalHeaderLabels(["Tên môn","Số lượng đề thi", "Số lượng bài chấm", "Tỉ lệ đề thi", "Tỉ lệ bài chấm"])
        self.bao_cao_nam.setColumnWidth(0, 100)
        self.bao_cao_nam.setColumnWidth(1, 150)
        self.bao_cao_nam.setColumnWidth(2, 150)
        self.bao_cao_nam.setColumnWidth(3, 150)
        self.bao_cao_nam.setColumnWidth(4, 150)
        
        self.bao_cao_nam.clearContents()
        self.bao_cao_nam.setRowCount(0)

        self.tong_so_bai_cham.clear()
        self.tong_so_de_thi.clear()